from openpyxl import Workbook
import openpyxl
import requests
import time
import os
import re

def create_workbook():
    """Create a new workbook with headers"""
    workbook = Workbook()
    sheet = workbook.active
    
    # Define headers
    headers = [
        'Business Name',
        'Street Address',
        'Postal Code',
        'City',
        'Phone',
        'Website',
        'Average Rating',
        'Number of Reviews',
        'Reviews'
    ]
    
    # Write headers
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col, value=header)
    
    return workbook

def write_data_row(sheet, row_num, data):
    """Write a row of data to the worksheet"""
    # Add 2 to row_num because row 1 is headers
    actual_row = row_num + 2
    
    # Split address into components if it exists
    address = data.get('address', '')
    street_address = ''
    postal_code = ''
    city = ''
    
    if address:
        # Try to split address into components
        # Swiss/German format: Street Name 123, 1234 City
        parts = address.split(',')
        if len(parts) >= 2:
            street_address = parts[0].strip()
            location_part = parts[1].strip()
            
            # Try to extract postal code and city
            postal_match = re.search(r'\b\d{4,5}\b', location_part)
            if postal_match:
                postal_code = postal_match.group(0)
                # City is everything after the postal code
                city_match = re.search(rf'{postal_code}\s+(.*)', location_part)
                if city_match:
                    city = city_match.group(1).strip()
            else:
                # If no postal code found, just use the whole part as city
                city = location_part
        else:
            # If no comma, try to extract postal code directly
            postal_match = re.search(r'\b\d{4,5}\b', address)
            if postal_match:
                postal_code = postal_match.group(0)
                # Split address at the postal code
                parts = re.split(rf'\s*{postal_code}\s*', address)
                if len(parts) >= 2:
                    street_address = parts[0].strip()
                    city = parts[1].strip()
                else:
                    street_address = address
            else:
                # If still no postal code found
                street_address = address
    
    # Write data to cells
    sheet.cell(row=actual_row, column=1, value=data.get('name', ''))
    sheet.cell(row=actual_row, column=2, value=street_address)
    sheet.cell(row=actual_row, column=3, value=postal_code)
    sheet.cell(row=actual_row, column=4, value=city)
    sheet.cell(row=actual_row, column=5, value=data.get('phone', ''))
    sheet.cell(row=actual_row, column=6, value=data.get('website', ''))
    sheet.cell(row=actual_row, column=7, value=data.get('rating', ''))
    sheet.cell(row=actual_row, column=8, value=data.get('review_count', ''))
    
    # Format reviews with new fields if present
    reviews = data.get('reviews', [])
    if reviews:
        reviews_text = '\n\n'.join([
            f"Rating: {r.get('rating', '')} stars\n"
            f"Date: {r.get('date', '')}\n"
            f"Comment: {r.get('text', '')}\n"
            f"Positive Points: {', '.join(r.get('positive_points', []))}\n"
            f"Negative Points: {', '.join(r.get('negative_points', []))}\n"
            f"Services: {', '.join(r.get('services', []))}"
            for r in reviews
        ])
        sheet.cell(row=actual_row, column=9, value=reviews_text)
    else:
        sheet.cell(row=actual_row, column=9, value='')

def write_to_workbook(data, row, col, workbook_name="ScrapedData_GoogleMaps.xlsx", worksheet_name="Sheet1"):
    # Validate paths
    if not os.path.abspath(workbook_name).startswith(os.path.abspath(os.getcwd())):
        raise ValueError("Invalid workbook path")
    
    # Validate row/column inputs
    if not (isinstance(row, int) and isinstance(col, int)):
        raise ValueError("Invalid row or column")
    
    # Open the Excel file
    workbook = openpyxl.load_workbook(workbook_name)
    try:
        # Select the first sheet
        sheet = workbook[worksheet_name]
        # Write data to specified row and column
        sheet.cell(row=row, column=col).value = data
    finally:
        workbook.save(workbook_name)

def count_entries_in_workbook(workbook_name="ScrapedData_GoogleMaps.xlsx"):
    # Open the Excel file
    workbook = openpyxl.load_workbook(workbook_name)
    try:
        # Select the first sheet
        sheet = workbook.worksheets[0]
        # Count the number of rows in the sheet
        num_rows = sheet.max_row
    finally:
        workbook.close()
        # return the result
        return num_rows

def extract_column_from_row(row, column, workbook_name="ScrapedData_GoogleMaps.xlsx"):
    # Open the Excel file
    workbook = openpyxl.load_workbook(workbook_name)
    try:
        # Select the first sheet
        sheet = workbook.worksheets[0]
        # Extract value from specified column
        column_value = sheet.cell(row=row, column=column).value
    finally:
        workbook.close()
        return column_value

def get_website_data(url):
    if not url.startswith(('http://', 'https://')):
        return None, None
    
    try:
        response = requests.get(
            url,
            allow_redirects=False,  # Don't follow redirects
            timeout=10,
            verify=True  # Verify SSL certificates
        )
        # Add rate limiting
        time.sleep(1)
        return response.text, None
    except Exception as e:
        return None, str(e)